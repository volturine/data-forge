import logging
import uuid
from dataclasses import dataclass
from datetime import UTC, datetime
from pathlib import Path

import polars as pl
from openpyxl import load_workbook
from openpyxl.utils.cell import range_boundaries
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession

from core.exceptions import DataSourceNotFoundError, DataSourceValidationError, FileError
from modules.compute.operations.datasource import load_datasource
from modules.datasource.models import DataSource
from modules.datasource.schemas import ColumnSchema, CSVOptions, DataSourceResponse, SchemaInfo

logger = logging.getLogger(__name__)


async def create_file_datasource(
    session: AsyncSession,
    name: str,
    file_path: str,
    file_type: str,
    options: dict | None = None,
    csv_options: CSVOptions | None = None,
    sheet_name: str | None = None,
    start_row: int | None = None,
    start_col: int | None = None,
    end_col: int | None = None,
    end_row: int | None = None,
    has_header: bool | None = None,
    table_name: str | None = None,
    named_range: str | None = None,
) -> DataSourceResponse:
    """Create a file-based datasource."""
    datasource_id = str(uuid.uuid4())

    config = {
        'file_path': file_path,
        'file_type': file_type,
        'options': options or {},
        'csv_options': csv_options.model_dump() if csv_options else None,
        'sheet_name': sheet_name,
        'start_row': start_row,
        'start_col': start_col,
        'end_col': end_col,
        'end_row': end_row,
        'has_header': has_header,
        'table_name': table_name,
        'named_range': named_range,
    }

    datasource = DataSource(
        id=datasource_id,
        name=name,
        source_type='file',
        config=config,
        created_at=datetime.now(UTC),
    )

    session.add(datasource)
    await session.commit()
    await session.refresh(datasource)

    logger.info(f'Created file datasource {datasource_id} ({name}) with file {file_path}')
    return DataSourceResponse.model_validate(datasource)


@dataclass
class ExcelPreviewResult:
    preview: list[list[str | None]]
    detected_end_row: int | None
    sheet_name: str
    start_row: int
    start_col: int
    end_col: int


def build_excel_preview(
    file_path: Path,
    sheet_name: str,
    start_row: int,
    start_col: int,
    end_col: int,
    has_header: bool,
    table_name: str | None = None,
    named_range: str | None = None,
    preview_rows: int = 100,
) -> ExcelPreviewResult:
    workbook = load_workbook(file_path, read_only=False, data_only=True)
    resolved = _resolve_excel_bounds(
        workbook,
        sheet_name,
        start_row,
        start_col,
        end_col,
        table_name,
        named_range,
    )
    sheet = workbook[resolved.sheet_name]

    detected_end_row = resolved.end_row
    if detected_end_row is None:
        detected_end_row = _detect_end_row(sheet, resolved.start_row, resolved.start_col, resolved.end_col)

    preview_end_row = min(resolved.start_row + preview_rows - 1, detected_end_row)
    rows = _collect_preview_rows(sheet, resolved.start_row, resolved.start_col, resolved.end_col, preview_end_row)
    return ExcelPreviewResult(
        preview=rows,
        detected_end_row=detected_end_row,
        sheet_name=resolved.sheet_name,
        start_row=resolved.start_row,
        start_col=resolved.start_col,
        end_col=resolved.end_col,
    )


def resolve_excel_selection(
    file_path: Path,
    sheet_name: str,
    start_row: int,
    start_col: int,
    end_col: int,
    table_name: str | None = None,
    named_range: str | None = None,
) -> tuple[str, int, int, int, int]:
    workbook = load_workbook(file_path, read_only=False, data_only=True)
    resolved = _resolve_excel_bounds(
        workbook,
        sheet_name,
        start_row,
        start_col,
        end_col,
        table_name,
        named_range,
    )
    sheet = workbook[resolved.sheet_name]
    end_row = resolved.end_row
    if end_row is None:
        end_row = _detect_end_row(sheet, resolved.start_row, resolved.start_col, resolved.end_col)
    return resolved.sheet_name, resolved.start_row, resolved.start_col, resolved.end_col, end_row


class _ExcelBounds:
    def __init__(
        self,
        sheet_name: str,
        start_row: int,
        start_col: int,
        end_col: int,
        end_row: int | None,
    ) -> None:
        self.sheet_name = sheet_name
        self.start_row = start_row
        self.start_col = start_col
        self.end_col = end_col
        self.end_row = end_row


def _resolve_excel_bounds(
    workbook,
    sheet_name: str,
    start_row: int,
    start_col: int,
    end_col: int,
    table_name: str | None,
    named_range: str | None,
) -> _ExcelBounds:
    if table_name:
        sheet = workbook[sheet_name]
        tables = getattr(sheet, 'tables', None)
        if not tables:
            raise ValueError(f'No tables available in sheet: {sheet_name}')
        table = tables.get(table_name)
        if not table:
            raise ValueError(f'Table not found: {table_name}')
        min_col, min_row, max_col, max_row = range_boundaries(table.ref)
        if min_col is None or min_row is None or max_col is None or max_row is None:
            raise ValueError(f'Invalid table range: {table_name}')
        min_col = int(min_col)
        min_row = int(min_row)
        max_col = int(max_col)
        max_row = int(max_row)
        return _ExcelBounds(sheet_name, min_row - 1, min_col - 1, max_col - 1, max_row - 1)

    if named_range:
        defined = workbook.defined_names.get(named_range)
        if not defined:
            raise ValueError(f'Named range not found: {named_range}')
        destinations = list(defined.destinations)
        if not destinations:
            raise ValueError(f'Named range has no destinations: {named_range}')
        dest_sheet, coord = destinations[0]
        min_col, min_row, max_col, max_row = range_boundaries(coord)
        if min_col is None or min_row is None or max_col is None or max_row is None:
            raise ValueError(f'Invalid named range: {named_range}')
        min_col = int(min_col)
        min_row = int(min_row)
        max_col = int(max_col)
        max_row = int(max_row)
        return _ExcelBounds(dest_sheet, min_row - 1, min_col - 1, max_col - 1, max_row - 1)

    return _ExcelBounds(sheet_name, max(start_row, 0), max(start_col, 0), max(end_col, start_col), None)


def _detect_end_row(sheet, start_row: int, start_col: int, end_col: int) -> int:
    max_row = sheet.max_row or 0
    if max_row <= start_row:
        return start_row
    for row_index in range(start_row + 1, max_row + 1):
        values = []
        for cell in sheet.iter_rows(min_row=row_index, max_row=row_index, min_col=start_col + 1, max_col=end_col + 1):
            values = [c.value for c in cell]
        if all(value is None or str(value).strip() == '' for value in values):
            return max(start_row, row_index - 2)
    return max_row - 1


def _collect_preview_rows(
    sheet,
    start_row: int,
    start_col: int,
    end_col: int,
    preview_end_row: int,
) -> list[list[str | None]]:
    rows: list[list[str | None]] = []
    for row in sheet.iter_rows(
        min_row=start_row + 1,
        max_row=preview_end_row + 1,
        min_col=start_col + 1,
        max_col=end_col + 1,
    ):
        values = [cell.value for cell in row]
        rows.append([str(value) if value is not None else None for value in values])
    return rows


async def create_database_datasource(
    session: AsyncSession,
    name: str,
    connection_string: str,
    query: str,
) -> DataSourceResponse:
    datasource_id = str(uuid.uuid4())

    config = {
        'connection_string': connection_string,
        'query': query,
    }

    datasource = DataSource(
        id=datasource_id,
        name=name,
        source_type='database',
        config=config,
        created_at=datetime.now(UTC),
    )

    session.add(datasource)
    await session.commit()
    await session.refresh(datasource)

    return DataSourceResponse.model_validate(datasource)


async def create_api_datasource(
    session: AsyncSession,
    name: str,
    url: str,
    method: str = 'GET',
    headers: dict | None = None,
    auth: dict | None = None,
) -> DataSourceResponse:
    datasource_id = str(uuid.uuid4())

    config = {
        'url': url,
        'method': method,
        'headers': headers,
        'auth': auth,
    }

    datasource = DataSource(
        id=datasource_id,
        name=name,
        source_type='api',
        config=config,
        created_at=datetime.now(UTC),
    )

    session.add(datasource)
    await session.commit()
    await session.refresh(datasource)

    return DataSourceResponse.model_validate(datasource)


async def create_duckdb_datasource(
    session: AsyncSession,
    name: str,
    db_path: str | None,
    query: str,
    read_only: bool = True,
) -> DataSourceResponse:
    """Create a DuckDB datasource."""
    datasource_id = str(uuid.uuid4())

    config = {
        'db_path': db_path,
        'query': query,
        'read_only': read_only,
    }

    datasource = DataSource(
        id=datasource_id,
        name=name,
        source_type='duckdb',
        config=config,
        created_at=datetime.now(UTC),
    )

    session.add(datasource)
    await session.commit()
    await session.refresh(datasource)

    logger.info(f'Created DuckDB datasource {datasource_id} ({name})')
    return DataSourceResponse.model_validate(datasource)


async def get_datasource_schema(session: AsyncSession, datasource_id: str, sheet_name: str | None = None) -> SchemaInfo:
    """Get or extract schema for a datasource."""
    result = await session.execute(select(DataSource).where(DataSource.id == datasource_id))
    datasource = result.scalar_one_or_none()

    if not datasource:
        raise DataSourceNotFoundError(datasource_id)

    # Check if we have cached schema with row_count
    if datasource.schema_cache and sheet_name is None:
        cached = SchemaInfo.model_validate(datasource.schema_cache)
        # If row_count is missing from cache, re-extract to get it
        if cached.row_count is not None:
            logger.debug(f'Using cached schema for datasource {datasource_id}')
            return cached

    logger.info(f'Extracting schema for datasource {datasource_id}')
    schema_info = await _extract_schema(datasource, sheet_name=sheet_name)

    if sheet_name is None:
        datasource.schema_cache = schema_info.model_dump()
        await session.commit()

    logger.info(f'Schema extracted and cached for datasource {datasource_id}: {len(schema_info.columns)} columns')
    return schema_info


async def _extract_schema(datasource: DataSource, sheet_name: str | None = None) -> SchemaInfo:
    if datasource.source_type == 'file':
        config = {
            'source_type': datasource.source_type,
            **datasource.config,
        }
        if sheet_name:
            config = {**config, 'sheet_name': sheet_name}
        lazy = load_datasource(config)
        schema = lazy.collect_schema()
        row_count = lazy.select(pl.len()).collect().item()
        sheet_names = None

        columns = [
            ColumnSchema(
                name=name,
                dtype=str(dtype),
                nullable=True,
            )
            for name, dtype in schema.items()
        ]

        return SchemaInfo(columns=columns, row_count=row_count, sheet_names=sheet_names)

    if datasource.source_type == 'database':
        connection_string = datasource.config['connection_string']
        query = datasource.config['query']

        frame = pl.read_database(query, connection_string)
        schema = frame.schema
        row_count = frame.height

        columns = [
            ColumnSchema(
                name=name,
                dtype=str(dtype),
                nullable=True,
            )
            for name, dtype in schema.items()
        ]

        return SchemaInfo(columns=columns, row_count=row_count)

    if datasource.source_type == 'duckdb':
        config = {
            'source_type': datasource.source_type,
            **datasource.config,
        }
        lazy = load_datasource(config)
        schema = lazy.collect_schema()
        row_count = lazy.select(pl.len()).collect().item()

        columns = [
            ColumnSchema(
                name=name,
                dtype=str(dtype),
                nullable=True,
            )
            for name, dtype in schema.items()
        ]

        return SchemaInfo(columns=columns, row_count=row_count)

    raise DataSourceValidationError(
        f'Schema extraction not supported for type: {datasource.source_type}',
        details={'source_type': datasource.source_type},
    )


async def get_datasource(session: AsyncSession, datasource_id: str) -> DataSourceResponse:
    result = await session.execute(select(DataSource).where(DataSource.id == datasource_id))
    datasource = result.scalar_one_or_none()

    if not datasource:
        raise DataSourceNotFoundError(datasource_id)

    return DataSourceResponse.model_validate(datasource)


async def list_datasources(session: AsyncSession) -> list[DataSourceResponse]:
    result = await session.execute(select(DataSource))
    datasources = result.scalars().all()
    return [DataSourceResponse.model_validate(ds) for ds in datasources]


async def delete_datasource(session: AsyncSession, datasource_id: str) -> None:
    """Delete a datasource and its associated file if it exists."""
    result = await session.execute(select(DataSource).where(DataSource.id == datasource_id))
    datasource = result.scalar_one_or_none()

    if not datasource:
        raise DataSourceNotFoundError(datasource_id)

    # Delete associated file if it's a file datasource
    if datasource.source_type == 'file' and 'file_path' in datasource.config:
        file_path = Path(datasource.config['file_path'])
        if file_path.exists():
            try:
                # Check if file is accessible before deletion
                if not file_path.is_file():
                    logger.warning(f'Path exists but is not a file: {file_path}')
                else:
                    file_path.unlink()
                    logger.info(f'Deleted file: {file_path}')
            except PermissionError as e:
                logger.error(f'Permission denied when deleting file {file_path}: {e}')
                raise FileError(
                    f'Permission denied when deleting file: {file_path}',
                    error_code='FILE_PERMISSION_DENIED',
                    details={'file_path': str(file_path)},
                )
            except OSError as e:
                logger.error(f'OS error when deleting file {file_path}: {e}')
                raise FileError(
                    f'Failed to delete file: {file_path}',
                    error_code='FILE_DELETE_ERROR',
                    details={'file_path': str(file_path), 'error': str(e)},
                )

    await session.delete(datasource)
    await session.commit()
    logger.info(f'Deleted datasource {datasource_id}')
