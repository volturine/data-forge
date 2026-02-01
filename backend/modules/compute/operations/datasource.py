from collections.abc import Callable

import polars as pl
from openpyxl import load_workbook
from pydantic import ConfigDict

from modules.compute.operations.base import OperationHandler, OperationParams


class DatasourceParams(OperationParams):
    model_config = ConfigDict(extra='allow')

    source_type: str = 'file'
    file_path: str | None = None
    file_type: str | None = None
    options: dict | None = None
    csv_options: dict | None = None
    sheet_name: str | None = None
    start_row: int | None = None
    start_col: int | None = None
    end_col: int | None = None
    end_row: int | None = None
    has_header: bool | None = None
    table_name: str | None = None
    named_range: str | None = None
    connection_string: str | None = None
    query: str | None = None
    db_path: str | None = None
    read_only: bool = True
    column_schema: list[dict] | None = None


class DatasourceHandler(OperationHandler):
    FILE_LOADERS: dict[str, Callable[[str, dict], pl.LazyFrame]] = {
        'csv': lambda path, opts: pl.scan_csv(path, **DatasourceHandler._csv_opts(opts)),
        'parquet': lambda path, _: pl.scan_parquet(path),
        'json': lambda path, _: pl.read_json(path).lazy(),
        'ndjson': lambda path, _: pl.scan_ndjson(path),
        'excel': lambda path, opts: DatasourceHandler._read_excel(path, opts),
    }

    @property
    def name(self) -> str:
        return 'datasource'

    def __call__(
        self,
        lf: pl.LazyFrame,
        params: dict,
        *,
        right_lf: pl.LazyFrame | None = None,
        right_sources: dict[str, pl.LazyFrame] | None = None,
    ) -> pl.LazyFrame:
        validated = DatasourceParams.model_validate(params)
        loaders = {
            'file': self._load_file,
            'database': self._load_database,
            'duckdb': self._load_duckdb,
        }
        loader = loaders.get(validated.source_type)
        if not loader:
            raise ValueError(f'Unsupported source type: {validated.source_type}')
        return loader(validated)

    @staticmethod
    def _csv_opts(opts: dict | None) -> dict:
        if not opts:
            return {}
        return {
            'separator': opts.get('delimiter', ','),
            'quote_char': opts.get('quote_char', '"'),
            'has_header': opts.get('has_header', True),
            'skip_rows': opts.get('skip_rows', 0),
            'encoding': opts.get('encoding', 'utf8'),
        }

    def _load_file(self, config: DatasourceParams) -> pl.LazyFrame:
        if not config.file_path or not config.file_type:
            raise ValueError('Datasource file loading requires file_path and file_type')
        if config.file_type == 'excel' and _has_bounds(config):
            return self._read_excel_bounds(config)
        opts = config.csv_options or config.options or {}
        opts = self._merge_excel_opts(config, opts)
        loader = self.FILE_LOADERS.get(config.file_type)
        if not loader:
            raise ValueError(f'Unsupported file type: {config.file_type}')
        return loader(config.file_path, opts)

    def _load_database(self, config: DatasourceParams) -> pl.LazyFrame:
        if not config.connection_string or not config.query:
            raise ValueError('Datasource database loading requires connection_string and query')
        return pl.read_database(config.query, config.connection_string).lazy()

    def _load_duckdb(self, config: DatasourceParams) -> pl.LazyFrame:
        import duckdb

        if not config.query:
            raise ValueError('Datasource DuckDB loading requires query')
        conn = (
            duckdb.connect(database=config.db_path, read_only=config.read_only) if config.db_path else duckdb.connect(database=':memory:')
        )
        try:
            return conn.execute(config.query).fetch_df().lazy()
        finally:
            conn.close()

    @staticmethod
    def _read_excel(path: str, opts: dict) -> pl.LazyFrame:
        sheet_name = opts.get('sheet_name')
        table_name = opts.get('table_name')
        named_range = opts.get('named_range')
        next_opts: dict = {}
        if sheet_name is not None:
            next_opts['sheet_name'] = sheet_name
        if table_name is not None:
            next_opts['table_name'] = table_name
        if named_range is not None:
            next_opts['named_range'] = named_range
        return pl.read_excel(path, **next_opts).lazy()

    @staticmethod
    def _merge_excel_opts(config: DatasourceParams, opts: dict) -> dict:
        next_opts = opts
        if config.sheet_name:
            next_opts = {**next_opts, 'sheet_name': config.sheet_name}
        if config.table_name:
            next_opts = {**next_opts, 'table_name': config.table_name}
        if config.named_range:
            next_opts = {**next_opts, 'named_range': config.named_range}
        if config.has_header is not None:
            next_opts = {**next_opts, 'has_header': config.has_header}
        return next_opts

    @staticmethod
    def _read_excel_bounds(config: DatasourceParams) -> pl.LazyFrame:
        file_path = config.file_path
        sheet_name = config.sheet_name
        start_row = config.start_row
        start_col = config.start_col
        end_row = config.end_row
        end_col = config.end_col
        has_header = config.has_header if config.has_header is not None else True
        if not file_path or start_row is None or start_col is None or end_row is None or end_col is None:
            raise ValueError('Excel bounds require file_path, start_row, start_col, end_row, end_col')

        workbook = load_workbook(file_path, read_only=True, data_only=True)
        target_sheet = sheet_name or workbook.sheetnames[0]
        sheet = workbook[target_sheet]
        rows = list(
            sheet.iter_rows(
                min_row=start_row + 1,
                max_row=end_row + 1,
                min_col=start_col + 1,
                max_col=end_col + 1,
                values_only=True,
            )
        )

        if not rows:
            return pl.DataFrame().lazy()

        if has_header:
            header = rows[0]
            columns = _normalize_headers(header)
            data_rows = rows[1:]
        else:
            columns = [f'column_{index + 1}' for index in range(len(rows[0]))]
            data_rows = rows
        frame = pl.DataFrame(data_rows, schema=columns)
        return frame.lazy()


def _normalize_headers(values: tuple[object | None, ...]) -> list[str]:
    names: list[str] = []
    seen: dict[str, int] = {}
    for index, value in enumerate(values):
        base = str(value).strip() if value is not None else f'column_{index + 1}'
        if base not in seen:
            seen[base] = 0
            names.append(base)
            continue
        seen[base] += 1
        names.append(f'{base}_{seen[base]}')
    return names


def _has_bounds(config: DatasourceParams) -> bool:
    return config.start_row is not None and config.start_col is not None and config.end_col is not None and config.end_row is not None


def load_datasource(config: dict) -> pl.LazyFrame:
    return DatasourceHandler()(pl.LazyFrame(), config)
