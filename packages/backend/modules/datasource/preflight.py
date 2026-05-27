import asyncio
import contextlib
import uuid
from dataclasses import dataclass
from datetime import UTC, datetime, timedelta
from pathlib import Path

from core.object_store import delete_object, is_object_store_url
from openpyxl import load_workbook


@dataclass
class ExcelPreflight:
    source_path: str
    sheets: list[str]
    tables: dict[str, list[str]]
    named_ranges: list[str]
    created_at: datetime
    delete_source: bool


_PREFLIGHTS: dict[str, ExcelPreflight] = {}
_PREFLIGHTS_LOCK = asyncio.Lock()
_PREFLIGHT_TTL = timedelta(minutes=30)


async def create_preflight(file_path: Path, *, source_path: str | None = None, delete_source: bool = True) -> tuple[str, ExcelPreflight]:
    workbook = await asyncio.to_thread(load_workbook, file_path, read_only=False, data_only=True)
    sheets = workbook.sheetnames
    tables: dict[str, list[str]] = {}
    for sheet in workbook.worksheets:
        if not hasattr(sheet, "tables") or not sheet.tables:
            continue
        tables[sheet.title] = list(sheet.tables.keys())

    named_ranges = [name for name in workbook.defined_names]
    preflight_id = str(uuid.uuid4())
    preflight = ExcelPreflight(
        source_path=source_path or str(file_path),
        sheets=sheets,
        tables=tables,
        named_ranges=named_ranges,
        created_at=datetime.now(UTC).replace(tzinfo=None),
        delete_source=delete_source,
    )
    async with _PREFLIGHTS_LOCK:
        _PREFLIGHTS[preflight_id] = preflight
    return preflight_id, preflight


async def get_preflight(preflight_id: str) -> ExcelPreflight | None:
    await _cleanup_expired()
    async with _PREFLIGHTS_LOCK:
        return _PREFLIGHTS.get(preflight_id)


async def clear_preflight(preflight_id: str, *, delete_source: bool = True) -> None:
    async with _PREFLIGHTS_LOCK:
        preflight = _PREFLIGHTS.pop(preflight_id, None)
        if not preflight:
            return
    await _delete_source(preflight.source_path, delete_source=delete_source and preflight.delete_source)


async def _cleanup_expired() -> None:
    now = datetime.now(UTC).replace(tzinfo=None)
    async with _PREFLIGHTS_LOCK:
        expired: list[ExcelPreflight] = []
        for preflight_id, preflight in list(_PREFLIGHTS.items()):
            if now - preflight.created_at <= _PREFLIGHT_TTL:
                continue
            expired.append(_PREFLIGHTS.pop(preflight_id))
    for preflight in expired:
        await _delete_source(preflight.source_path, delete_source=preflight.delete_source)


async def _delete_source(path: str, *, delete_source: bool) -> None:
    if not delete_source:
        return
    if is_object_store_url(path):
        await asyncio.to_thread(delete_object, path)
        return
    local_path = Path(path)
    if not local_path.exists():
        return
    with contextlib.suppress(FileNotFoundError):
        await asyncio.to_thread(local_path.unlink)
